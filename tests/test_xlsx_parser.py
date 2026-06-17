import io

import openpyxl
import pytest

from ai_employee.ingestion_worker.parsers import XlsxParser, ParsedSection


@pytest.fixture
def xlsx_bytes() -> bytes:
    """Generate a minimal valid XLSX in memory with headers and data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alarms"

    # Header row
    ws.append(["Time", "Event", "Severity", "Node"])

    # Data rows
    ws.append(["10:00", "Link Down", "Critical", "BSC-01"])
    ws.append(["10:05", "Failover", "Warning", "RNC-02"])
    ws.append(["10:15", "Link Up", "Info", "BSC-01"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def xlsx_multisheet_bytes() -> bytes:
    """Generate an XLSX with multiple sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Alarms"
    ws1.append(["Time", "Event"])
    ws1.append(["10:00", "Link Down"])

    ws2 = wb.create_sheet("KPIs")
    ws2.append(["Metric", "Value"])
    ws2.append(["Availability", "99.99%"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestXlsxParser:
    def test_sheet_name_as_section_path(self, xlsx_bytes: bytes) -> None:
        """section_path should be the sheet name."""
        sections = XlsxParser().parse(xlsx_bytes)
        assert len(sections) >= 1
        assert sections[0].section_path == "Alarms"

    def test_data_rows_become_blocks(self, xlsx_bytes: bytes) -> None:
        """Each data row should become a block with col_name: value format."""
        sections = XlsxParser().parse(xlsx_bytes)
        # Only data blocks (not header-only)
        blocks = sections[0].blocks
        assert len(blocks) == 3  # 3 data rows
        assert "Link Down" in blocks[0]
        assert "Critical" in blocks[0]
        assert "Failover" in blocks[1]
        assert "BSC-01" in blocks[2]

    def test_header_preserved_in_output(self, xlsx_bytes: bytes) -> None:
        """Column names from header row should appear in each block."""
        sections = XlsxParser().parse(xlsx_bytes)
        first_block = sections[0].blocks[0]
        assert "Time" in first_block
        assert "Event" in first_block
        assert "Severity" in first_block
        assert "Node" in first_block

    def test_multiple_sheets(self, xlsx_multisheet_bytes: bytes) -> None:
        """Multi-sheet workbooks should produce a section per sheet."""
        sections = XlsxParser().parse(xlsx_multisheet_bytes)
        paths = {s.section_path for s in sections}
        assert "Alarms" in paths
        assert "KPIs" in paths

    def test_empty_sheet_no_data(self) -> None:
        """Sheet with only headers and no data rows should have no blocks."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        ws.append(["Col1", "Col2"])
        buf = io.BytesIO()
        wb.save(buf)
        sections = XlsxParser().parse(buf.getvalue())
        if sections:
            assert sections[0].blocks == []

    def test_empty_workbook(self) -> None:
        """Empty workbook with no data should return empty list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        buf = io.BytesIO()
        wb.save(buf)
        sections = XlsxParser().parse(buf.getvalue())
        assert sections == [] or all(len(s.blocks) == 0 for s in sections)
