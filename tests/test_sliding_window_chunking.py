"""Sliding-window chunking + table structure tests."""
from __future__ import annotations

import io

from ai_employee.ingestion_worker.chunker import (
    sliding_window_chunk,
)
from ai_employee.ingestion_worker.parsers import XlsxParser


def _make_xlsx(rows: list[list[str]]) -> bytes:
    """Build a minimal xlsx in-memory using openpyxl."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI_Sheet"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# --------------------------------------------------------------------------- #
# sliding_window_chunk
# --------------------------------------------------------------------------- #


def test_sliding_window_chunk_short_text_returns_unchanged() -> None:
    chunks = sliding_window_chunk("hello world", window_size=100, overlap=10)
    assert chunks == ["hello world"]


def test_sliding_window_chunk_splits_long_text() -> None:
    text = "a" * 250
    chunks = sliding_window_chunk(text, window_size=100, overlap=20)
    assert len(chunks) >= 3
    assert all(len(c) <= 100 for c in chunks)


def test_sliding_window_chunk_has_overlap() -> None:
    text = "abcdefghij" * 20  # 200 chars
    chunks = sliding_window_chunk(text, window_size=60, overlap=20)
    assert len(chunks) >= 2
    # Adjacent chunks should share some trailing/leading characters.
    head_of_first = chunks[0][-20:]
    assert head_of_first in chunks[1]


def test_sliding_window_chunk_handles_empty_string() -> None:
    assert sliding_window_chunk("", window_size=100, overlap=10) == []


def test_sliding_window_chunk_respects_sentence_boundary_when_possible() -> None:
    text = "第一句。" + ("字" * 80) + "。第二句。" + ("字" * 80) + "。结束。"
    chunks = sliding_window_chunk(text, window_size=40, overlap=10)
    # All chunks should be non-empty and respect the size limit (with sentence cut).
    assert all(len(c) <= 50 for c in chunks)
    assert all(c.strip() for c in chunks)


def test_sliding_window_chunk_advances_by_window_minus_overlap() -> None:
    text = "x" * 1000
    chunks = sliding_window_chunk(text, window_size=100, overlap=25)
    # Each new chunk starts at most window_size chars after previous start.
    # Expected advances: 75, 75, ... → roughly 1000/75 = 13-14 chunks.
    assert 10 <= len(chunks) <= 20


# --------------------------------------------------------------------------- #
# XlsxParser table structure
# --------------------------------------------------------------------------- #


def test_xlsx_parser_emits_table_id_and_row_id_in_blocks() -> None:
    data = _make_xlsx(
        [
            ["site", "kpi", "value"],
            ["BJ-001", "PRB", "12.3"],
            ["BJ-001", "RRC", "9.1"],
            ["SH-002", "PRB", "8.4"],
        ]
    )
    parser = XlsxParser()
    sections = parser.parse(data)
    assert len(sections) == 1
    sec = sections[0]
    assert sec.table_id == "KPI_Sheet"
    assert sec.section_path == "KPI_Sheet"
    # Each data row should be exposed with a row_id.
    assert len(sec.row_ids) == 3
    assert sec.row_ids[0] == "row_001"
    assert sec.row_ids[1] == "row_002"
    assert sec.row_ids[2] == "row_003"
    # Blocks contain the original col:val text.
    assert "site: BJ-001" in sec.blocks[0]
    assert "PRB" in sec.blocks[0]


def test_xlsx_parser_handles_multiple_sheets() -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Alarms"
    ws1.append(["code", "site"])
    ws1.append(["AL-12", "BJ-001"])
    ws2 = wb.create_sheet("KPIs")
    ws2.append(["kpi", "value"])
    ws2.append(["PRB", "10.0"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()

    sections = XlsxParser().parse(data)
    assert [s.table_id for s in sections] == ["Alarms", "KPIs"]
    assert all(s.row_ids for s in sections)


def test_xlsx_parser_empty_sheet_skipped() -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    ws.append(["only_header"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    sections = XlsxParser().parse(data)
    # No data rows → section is skipped.
    assert sections == []


def test_xlsx_parser_rows_indexed_from_one() -> None:
    data = _make_xlsx([["h"], ["a"], ["b"], ["c"]])
    sections = XlsxParser().parse(data)
    assert sections[0].row_ids == ["row_001", "row_002", "row_003"]
