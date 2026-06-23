from __future__ import annotations

import io as _io
from dataclasses import dataclass, field

from bs4 import BeautifulSoup
from markdown_it import MarkdownIt


@dataclass
class ParsedSection:
    section_path: str
    blocks: list[str] = field(default_factory=list)
    # Table structure (xlsx / csv-style structured docs).  Empty for plain
    # prose sources.  ``table_id`` identifies the source table (typically
    # the sheet name); ``row_ids`` are 1-based row identifiers parallel
    # to ``blocks`` (``row_ids[i]`` is the row that produced
    # ``blocks[i]``).
    table_id: str | None = None
    row_ids: list[str] = field(default_factory=list)


class _BaseParser:
    def parse(self, source: str | bytes) -> list[ParsedSection]:
        raise NotImplementedError


class TextParser(_BaseParser):
    """纯文本：按空行分段，section_path 固定 root。"""

    def parse(self, source: str) -> list[ParsedSection]:
        blocks = [b.strip() for b in source.split("\n\n") if b.strip()]
        if not blocks:
            return []
        return [ParsedSection(section_path="root", blocks=blocks)]


class MarkdownParser(_BaseParser):
    """Markdown：按标题层级构造 section_path，正文段落作为 block。"""

    def parse(self, source: str) -> list[ParsedSection]:
        md = MarkdownIt()
        tokens = md.parse(source)
        sections: list[ParsedSection] = []
        heading_stack: list[str] = []
        inline_buffer: list[str] = []
        in_heading = False

        def current_path() -> str:
            return " > ".join(heading_stack) if heading_stack else "root"

        def flush_buffer() -> None:
            if not inline_buffer:
                return
            text = " ".join(inline_buffer).strip()
            inline_buffer.clear()
            if not text:
                return
            if sections and sections[-1].section_path == current_path():
                sections[-1].blocks.append(text)
            else:
                sections.append(ParsedSection(section_path=current_path(), blocks=[text]))

        for token in tokens:
            ttype = token.type
            if ttype == "heading_open":
                flush_buffer()
                in_heading = True
            elif ttype == "heading_close":
                heading_text = " ".join(inline_buffer).strip()
                inline_buffer.clear()
                if heading_text:
                    level = int(token.tag[1])
                    heading_stack = heading_stack[: level - 1]
                    heading_stack.append(heading_text)
                in_heading = False
            elif ttype == "inline":
                if token.content.strip():
                    inline_buffer.append(token.content.strip())
            elif ttype in {"paragraph_close", "bullet_list_close", "ordered_list_close"}:
                if not in_heading:
                    flush_buffer()

        flush_buffer()
        return sections


class HtmlParser(_BaseParser):
    """HTML：按 h1/h2/h3 切片，去除标签。"""

    _HEADING_TAGS = {"h1", "h2", "h3"}

    def parse(self, source: str) -> list[ParsedSection]:
        soup = BeautifulSoup(source, "html.parser")
        sections: list[ParsedSection] = []
        heading_stack: list[str] = []

        def current_path() -> str:
            return " > ".join(heading_stack) if heading_stack else "root"

        def append_block(text: str) -> None:
            text = text.strip()
            if not text:
                return
            if sections and sections[-1].section_path == current_path():
                sections[-1].blocks.append(text)
            else:
                sections.append(ParsedSection(section_path=current_path(), blocks=[text]))

        for el in soup.find_all(["h1", "h2", "h3", "p", "li"]):
            if el.name in self._HEADING_TAGS:
                heading_text = el.get_text(strip=True)
                if not heading_text:
                    continue
                level = int(el.name[1])
                heading_stack = heading_stack[: level - 1]
                heading_stack.append(heading_text)
            else:
                text = el.get_text(separator="", strip=True)
                if text:
                    append_block(text)

        return sections


class PdfParser(_BaseParser):
    """PDF：使用 pymupdf 提取文本，按页分组，section_path 为 "Page {n}"。

    扫描件（图片层无文本）走 OCR 兜底（spec §5.2）：当某页提取文本为空且
    注入了 ``ocr_backend`` 时，对该页渲染图后 OCR。OCR 不可用时静默跳过。
    """

    def __init__(self, ocr_backend: object | None = None) -> None:
        self._ocr_backend = ocr_backend

    def _extract_text(self, source: bytes) -> list[tuple[int, str]]:
        """Return [(page_num, page_text)] from a PDF's text layer."""
        import fitz

        doc = fitz.open(stream=source, filetype="pdf")
        try:
            out: list[tuple[int, str]] = []
            for page_num in range(1, len(doc) + 1):
                page = doc[page_num - 1]
                out.append((page_num, page.get_text("text")))
            return out
        finally:
            doc.close()

    def parse(self, source: str | bytes) -> list[ParsedSection]:
        if isinstance(source, str):
            data = source.encode("utf-8")
        else:
            data = source

        sections: list[ParsedSection] = []
        page_texts = self._extract_text(data)
        for page_num, text in page_texts:
            blocks = [b.strip() for b in text.split("\n\n") if b.strip()]
            if not blocks and text.strip():
                blocks = [text.strip()]
            # OCR fallback for image-only pages (spec §5.2 扫描件).
            if not blocks and self._ocr_backend is not None:
                ocr_text = self._ocr_page(data, page_num)
                if ocr_text:
                    blocks = [b.strip() for b in ocr_text.splitlines() if b.strip()]
            if blocks:
                sections.append(
                    ParsedSection(
                        section_path=f"Page {page_num}",
                        blocks=blocks,
                    )
                )
        return sections

    def _ocr_page(self, data: bytes, page_num: int) -> str:
        """Render one PDF page to an image and OCR it; '' on any failure."""
        if self._ocr_backend is None:
            return ""
        try:
            import fitz

            doc = fitz.open(stream=data, filetype="pdf")
            try:
                page = doc[page_num - 1]
                pix = page.get_pixmap(dpi=200)
                img_bytes = pix.tobytes("png")
            finally:
                doc.close()
        except Exception:
            return ""
        try:
            result = self._ocr_backend.ocr(img_bytes)  # type: ignore[attr-defined]
            return getattr(result, "text", "") or ""
        except Exception:
            return ""


class DocxParser(_BaseParser):
    """DOCX：使用 python-docx 按段落遍历，heading 样式构建 section_path，正文段落作为 blocks。"""

    def parse(self, source: str | bytes) -> list[ParsedSection]:
        import docx as _docx

        if isinstance(source, str):
            data = source.encode("utf-8")
        else:
            data = source

        doc = _docx.Document(_io.BytesIO(data))
        sections: list[ParsedSection] = []
        heading_stack: list[str] = []

        def current_path() -> str:
            return " > ".join(heading_stack) if heading_stack else "root"

        def append_block(text: str) -> None:
            text = text.strip()
            if not text:
                return
            if sections and sections[-1].section_path == current_path():
                sections[-1].blocks.append(text)
            else:
                sections.append(ParsedSection(section_path=current_path(), blocks=[text]))

        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue

            style_name = para.style.name if para.style else ""
            # Detect heading styles: "Heading 1", "Heading 2", "Heading 3", etc.
            if style_name.startswith("Heading ") and style_name[8:].isdigit():
                try:
                    level = int(style_name[8:])
                except ValueError:
                    level = 1
                heading_stack = heading_stack[: level - 1]
                heading_stack.append(text)
            else:
                append_block(text)

        return sections


class XlsxParser(_BaseParser):
    """XLSX：使用 openpyxl 读取每个 sheet，首行为表头，其余行转 "col: value" 文本块。"""

    def parse(self, source: str | bytes) -> list[ParsedSection]:
        import openpyxl

        if isinstance(source, str):
            data = source.encode("utf-8")
        else:
            data = source

        wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True)
        sections: list[ParsedSection] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(h) if h is not None else "" for h in rows[0]]
                data_rows = rows[1:]
                if not data_rows:
                    # sheet has only headers, no data blocks
                    continue

                blocks: list[str] = []
                row_ids: list[str] = []
                for row_idx, row in enumerate(data_rows, start=1):
                    parts: list[str] = []
                    for j, val in enumerate(row):
                        col_name = headers[j] if j < len(headers) else f"Col{j}"
                        val_str = str(val) if val is not None else ""
                        parts.append(f"{col_name}: {val_str}")
                    block = " | ".join(parts)
                    if block.strip():
                        blocks.append(block)
                        row_ids.append(f"row_{row_idx:03d}")

                if blocks:
                    sections.append(
                        ParsedSection(
                            section_path=sheet_name,
                            blocks=blocks,
                            table_id=sheet_name,
                            row_ids=row_ids,
                        )
                    )
        finally:
            wb.close()

        return sections


class ImageParser(_BaseParser):
    """图片（PNG/JPEG）：走 OCR backend 提取文本，section_path 固定 ``ocr``。

    与 :class:`OcrParser` 的区别：当 OCR backend 不可用（``OCR_BACKEND=disabled``
    默认）时，本 parser 不返回空列表，而是返回一个带占位文本的 section，
    使图片文档不会因 OCR 缺失而抛错——上层仍可入库（占位 chunk），后续可补 OCR。
    """

    _DISABLED_PLACEHOLDER = "[ocr unavailable: image not OCRed]"

    def __init__(self, ocr_backend: object | None = None) -> None:
        if ocr_backend is None:
            # Local import avoids a circular import at module load time
            # (ocr.py imports ParsedSection from parsers.py).
            from ai_employee.ingestion_worker.ocr import build_ocr_backend

            ocr_backend = build_ocr_backend()
        self._ocr_backend = ocr_backend

    def parse(self, source: str | bytes) -> list[ParsedSection]:
        if isinstance(source, str):
            data = source.encode("utf-8")
        else:
            data = source

        available = bool(getattr(self._ocr_backend, "available", True))
        if not available:
            # Graceful degradation: placeholder section, never crash.
            return [ParsedSection(section_path="ocr", blocks=[self._DISABLED_PLACEHOLDER])]

        try:
            result = self._ocr_backend.ocr(data)  # type: ignore[attr-defined]
        except Exception:
            return [ParsedSection(section_path="ocr", blocks=[self._DISABLED_PLACEHOLDER])]

        text = getattr(result, "text", "") or ""
        blocks = [line.strip() for line in text.splitlines() if line.strip()]
        if not blocks:
            # Backend available but produced no text → placeholder, not empty.
            return [ParsedSection(section_path="ocr", blocks=[self._DISABLED_PLACEHOLDER])]
        return [ParsedSection(section_path="ocr", blocks=blocks)]


class NotImplementedParser(_BaseParser):
    """占位解析器：明确返回不支持，不静默失败。"""

    def __init__(self, mime_type: str) -> None:
        self.mime_type = mime_type

    def parse(self, source: str | bytes) -> list[ParsedSection]:
        raise NotImplementedError(f"mime_unsupported: {self.mime_type}")


_PARSER_MAP = {
    "text/markdown": MarkdownParser,
    "text/html": HtmlParser,
    "text/plain": TextParser,
    "application/pdf": PdfParser,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": DocxParser,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": XlsxParser,
    "image/png": ImageParser,
    "image/jpeg": ImageParser,
}


def get_parser(mime_type: str) -> _BaseParser:
    cls = _PARSER_MAP.get(mime_type)
    if cls is None:
        return NotImplementedParser(mime_type)
    return cls()
